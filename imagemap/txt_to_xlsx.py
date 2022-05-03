import openpyxl

#write_excel(name,time) name放國家名稱，time放當日日期，此function會將該國家疫情資訊寫入至excel做統整
def write_excel(time):
    
    # 建立一個工作簿
    wb = openpyxl.Workbook()
    f = open('./image_map_txt/'+time+".txt",encoding="utf-8")

    # 建立一個表單
    sheet = wb.create_sheet("imagemap")
    sheet.cell(row=1, column=1, value="今日新增最多國家")
    sheet.cell(row=1, column=2, value="累計染疫人數最多國家")
    sheet.cell(row=1, column=3, value="近7日新增最快國家")
    sheet.cell(row=1, column=4, value="近7日下降最快國家")

    count=2     #紀錄row
    col=1       #紀錄col
    for line in f.readlines():
        sheet.cell(row=count, column=col, value=line)
        col+=1
         
    f.close()

    # 儲存
    wb.save('./image_map_xlsx/'+time+'.xlsx')