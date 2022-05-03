import openpyxl

#write_excel(name,time) name放國家名稱，time放當日日期，此function會將該國家疫情資訊寫入至excel做統整
def write_excel(name,time):
    
    # 建立一個工作簿
    workbook = openpyxl.load_workbook('./situation_xlsx/'+time+'.xlsx')
    f = open('./situation_txt/'+name+".txt",encoding="utf-8")

    # 建立一個表單
    sheet = workbook.create_sheet(name)
    sheet.cell(row=1, column=1, value="name")
    sheet.cell(row=1, column=2, value="sum")
    sheet.cell(row=1, column=3, value="day_new")
    sheet.cell(row=1, column=4, value="million")
    sheet.cell(row=1, column=5, value="death")
    sheet.cell(row=1, column=6, value="vaccine")

    count=2     #紀錄row
    col=1       #紀錄col
    for line in f.readlines():
        # 因為每一列資料的第4筆是空白所以寫入excel時要跳過
        if col==4:
            col+=1
            continue
        if col<4:
            sheet.cell(row=count, column=col, value=line)
            col+=1
        if col>4:
            sheet.cell(row=count, column=col-1, value=line)
            col+=1
        if count==2:    #因為只有全國有疫苗人數，所以全國那一行有其他都跳過
            if col==8:
                count+=1
                col=1
        else:
            if col==7:
                count+=1
                col=1
    f.close()

    # 儲存
    workbook.save('./situation_xlsx/'+time+'.xlsx')