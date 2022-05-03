import openpyxl

#write_excel(name,time) name放國家名稱，time放當日日期，此function會將該國家疫情資訊寫入至excel做統整

def write_excel(name,time):
    
    # 建立一個工作簿
    #workbook = openpyxl.Workbook()
    workbook = openpyxl.load_workbook("./news_xlsx/"+time+' news'+'.xlsx')
    f = open("./news_txt/"+ name +".txt",encoding="utf-8")

    # 建立一個表單
    sheet = workbook.create_sheet(name)
    sheet.cell(row=1, column=1, value="title")
    sheet.cell(row=1, column=2, value="content")
    sheet.cell(row=1, column=3, value="date_time")
    sheet.cell(row=1, column=4, value="url") 
    sheet.cell(row=1, column=5, value="image")
    sheet.cell(row=1, column=6, value="country") 
    sheet.cell(row=1, column=7, value="update_time")

    count = 2     #紀錄row
    col = 0       #紀錄col
    num=0

    for line in f.readlines():
        if (col>6):
            col=1
            count+=1
        else:
            col+=1
        sheet.cell(row=count, column=col, value=line)
    f.close()
    


    # for line in f.readlines():
    #     count = count + (num-1)/4
    #     sheet.cell(row=count, column=((num+1) % 4), value=line)
    
   
           


    # 儲存
    workbook.save("./news_xlsx/"+time+' news'+'.xlsx')