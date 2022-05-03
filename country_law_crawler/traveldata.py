import urllib.request as req
import bs4
""" from flask.wrappers import Response """
import openpyxl

url = "https://www.china-airlines.com/nz/zh/discover/news/travel-advisory/Immigration-Information"
request = req.Request(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"})
with req.urlopen(request) as response:
    data = response.read().decode("utf-8")

root = bs4.BeautifulSoup(data, "html.parser")

#篩出tbody與更新時間
data_tbody = root.find_all('tbody')
# print(data_tbody)
update_time = root.find(name = 'p', attrs = {"style":"text-align: right;"})
#print(update_time.get_text()) # 測試更新時間


#愈篩選出之國家
all_country=[' 日本',' 南韓',' 新加坡',' 馬來西亞',' 越南',' 泰國',' 菲律賓',' 印尼']

#用來存爬下來之"目的地"與"入境限制及檢疫規定"
data_country=[]
data_country_law=[]

#篩出tbody中每一個tr，存至data_tr(data_tr為一個二為陣列)
for tr in data_tbody:
    data_tr = tr.find_all('tr')


#篩出tr中每一個td，並依照style分出"目的地"與"入境限制及檢疫規定"，分別存至data_country(data_country為一個二為陣列)、data_country_law(data_country_law為一個二為陣列)
for td in data_tr:
    data_country.append(
        td.find_all(name = 'td', attrs = {"style":"width: 15%; text-align: center; vertical-align: top;"})
    )
    data_country_law.append(
        td.find_all(name = 'td', attrs = {"style":"width: 75%;"})
    )
# for i in data_country_law:
#     print(i)

#篩選出之條件
#日本','南韓','新加坡','馬來西亞','越南','泰國','菲律賓','印尼'
all_condition=['豁免條件 :','豁免條件：','檢疫規定','檢疫措施','備註','檢疫要求','檢疫條件 :','其他:','入境所須文件','官方網頁最新資訊','轉機限制','其他相關資訊']

#利用for迴圈比對data_country中國家那些是我們要的，我現在是print出來你可以執行看結果，爬下來的結果再看你要怎麼處理(補：中華航空沒有澳門)
wb = openpyxl.Workbook()
# 建立"更新時間(updatetime)"資料表
sheet = wb.create_sheet("updatetime", 1)
sheet.cell(row=1,column=1,value=update_time.get_text())

for i in range(1,len(data_country)):
    for country in all_country:
        # 若為目標國家
        if data_country[i][0].get_text() == country:
            print(data_country[i][0].get_text())  # 找出國家
            print(data_country_law[i][0].get_text()) # 找出國家法規的內文
            # 建立該國家的資料表
            ## sheet = wb.create_sheet(country, i-1)
            sheet = wb.create_sheet(country, i)
            # 第一欄放國名
            sheet.cell(row=1,column=1,value=country)
            
            # 找出該國符合"篩選條件"段落的首字位置(index值)，並放入regu_num[]中
            regu_num = [0]  # 注意：regu_num[]只有紀錄index值，內文存放在law中
            for condition in all_condition:
                law=data_country_law[i][0].get_text()   # 找到內文中符合"篩選條件"的段落並放入law中
                try:
                    regu_num.append(law.index(condition))   #接著記錄下該段落首字的位置(index值)
                except:
                    continue
            # 將"篩選條件"的index值從小到大排序    
            regu_num.sort()
           
            # 取出各個"篩選條件"的內文 (取出兩個index之間的段落內文)
            for i in range(0,len(regu_num)):
                if i == len(regu_num)-1:    # 判斷是否為最後一個段落
                    content=law[regu_num[i]:]
                else:
                    content=law[regu_num[i]:regu_num[i+1]]  # 取出兩個index之間的段落內文
                a = i+2
                sheet.cell(row=1,column=a,value=content) # 放入資料表中

        else:
            continue

#寫入澳門(長榮航空提供的資訊)
url = "https://www.evaair.com/zh-tw/customer-services/covid-19-information-center/travel-restrictions/?countryCode=MO"
request = req.Request(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"})
with req.urlopen(request) as response:
    data = response.read().decode("utf-8")

root = bs4.BeautifulSoup(data, "html.parser")

#篩出特定class
data_class = root.find_all('div',class_="editor editorC-editor")

#找取li
data_li=[]
for li in data_class:
    data_li.append(li.find_all('li'))

#載入資料表
sheet = wb.create_sheet("澳門")
sheet.cell(row=1,column=1,value="澳門")
#找取li的內容並寫入資料表
for i in range(0,len(data_li[0])):
    #print(data_li[0][i].get_text()) 
    sheet.cell(row=1,column=i+2,value=data_li[0][i].get_text())


#先載入sheet
sheet = wb['updatetime']
updatetime = sheet.cell(row=1, column=1).value
#儲存資料表
wb.save("CountryCon.xlsx")

