import openpyxl

def openWB(select_country):

    #判斷條件是否為Con(僅豁免和檢疫/長榮澳門：入境規定)
    if select_country[-3:] == "Con": #japanCon
        selected = 2
        select_country = select_country.replace('Con','')
    else: #japan
        selected = 1


    # 開啟工作簿
    workbook = openpyxl.load_workbook('../country_crawler/CountryCon.xlsx')
    #先載入sheet
    sheet = workbook[' 印尼']
    # 兩組做中英對照
    all_country_eng=['japan','korea','philippines','malaysia','singapore','thailand','vietnam','indonesia','macau']
    all_country=[' 日本',' 南韓',' 菲律賓',' 馬來西亞',' 新加坡',' 泰國',' 越南',' 印尼','澳門']
    
    for country_eng in all_country_eng: # country_eng表英文國家名
        if country_eng == select_country: # 當英文國名與選擇的英文相同時，取得其index值，並在相對應的中文國名字arr中，取得中文國名
            country_index=all_country_eng.index(country_eng)
            sheet_country =all_country[country_index]
            # 獲取表單(以中文國名)
            sheet = workbook[sheet_country] #注意：excel資料表為中文喔！

    # 讀取指定的單元格資料
    #cell = sheet.cell(row=1, column=2).value
    # 迴圈10次依序取出每格的資料，判斷內容不為空時印出，內容為空時則跳出迴圈
    count = 0
    for i in range(1,10):
        cell = sheet.cell(row=1, column=i).value 
        try:
            if i == 1:
                cells = "" 
            elif len(cell) != 0:
                ####我全部都有
                if selected == 1:
                    count = count + 1
                    cells = cells + cell 
                #####我只有豁免和檢疫/長榮澳門(入境規定)
                else:
                    count = count + 1
                    if cell[0:2] == "豁免" or cell[0:2] == "檢疫" or cell[0:4] == "入境規定":
                        cells += cell

        except:
            break

    return cells

def updatetime():
    # 開啟工作簿
    workbook = openpyxl.load_workbook('../country_crawler/CountryCon.xlsx')
    #先載入sheet
    sheet = workbook['updatetime']
    updatetime = sheet.cell(row=1, column=1).value
    return updatetime