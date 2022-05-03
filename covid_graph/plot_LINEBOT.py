import datetime
import xlrd
import openpyxl
import pyimgur
import matplotlib.pyplot as plt
import pgadmin4 as pg

''' 上傳圖片至imgur，回傳圖片url連結 '''

def upload_to_imgur(path):

    CLIENT_ID = "2f3efd9148b1935"
    PATH = path

    im = pyimgur.Imgur(CLIENT_ID)
    uploaded_image = im.upload_image(PATH, title="Uploaded with PyImgur")

    return uploaded_image.link  # 回傳url連結


'''變數設定'''

# 國家
country = {1: 'indonesia', 2: 'japan', 3: 'korea', 4: 'macao', 5: 'malaysia',6: 'philippines', 7: 'singapore', 8: 'vietnam', 9: 'thailand', 10: 'taiwan'}
country_chinese = {1: '印尼', 2: '日本', 3: '韓國', 4: '澳門',5: '馬來西亞', 6: '菲律賓', 7: '新加坡', 8: '越南', 9: '泰國', 10: '台灣'}
country_DAY7_url = {}
country_DAY30_url = {}
# excel檔名
savePLOTxlsx_name = './graph_xlsx/'+datetime.datetime.today().strftime("%Y-%m-%d")+" graph.xlsx"  # 存EXCEL檔名

# 檔案位置
PLOT7_path = './image/plot7.png'
PLOT30_path = './image/plot30.png'

'''!!!這邊改疫情資料放置位置!!!'''
COVID_data_path = pg.situation_save_path


'''跑所有國家'''
for j in country:  # 跑所有國家

    ''' DAY 7 圖表製作 '''
    # x軸的標
    time7_x = []  # 7天 => 在計算日期的時候直接存進去

    # Y軸數值
    day7_new_number = []  # 存每日新增 # 7天 => 每天數值都存進去(會有7筆資料)

    # 7天，包括今天

    # 直接取今天時間
    today = datetime.datetime.today()

    '''算日期'''
    # 用迴圈跑7次，倒算日期
    for i in range(7):

        time7_x.insert(0, str(today.strftime("%m-%d")))  # 插入list的最前面，因為是倒算

        today -= datetime.timedelta(days=1)  # 減一天，最後一次會減到，但是不會被存進去

    '''取數值，取7次'''
    for i in time7_x:
        date = i
        # 確認是否有檔案
        try:  # 有，即進檔案抓資料
            

            file_position = COVID_data_path+'2021-'+date+'.xlsx'  # 轉檔案名字
            print(file_position)
            workbook = xlrd.open_workbook(file_position)  # 開檔案

            # 選取sheet
            booksheet = workbook.sheet_by_name(country[j])  # 或用名稱取sheet
            
            #取daynew值 ###要換取其他值改這裡###
            print(type(booksheet.cell_value(1, 2)))
            print(booksheet.cell_value(1, 2))
            sum_number = int(booksheet.cell_value(1, 2).replace(',', ''))  # 取day_new的值，去除分位號(replace)，轉數值

            # 加入y軸串列(直接加在最後)
            day7_new_number.append(sum_number)
        except FileNotFoundError:
            sum_number = 0
            day7_new_number.append(sum_number)

    # 7天製圖表
    plt.plot(time7_x, day7_new_number)

    plt.title(country_chinese[j], weight='bold', size=16)  # 直接取國家名，第一個字母大寫
    plt.xlabel('日 期', weight='normal', size=12, labelpad=6)
    plt.ylabel('新 增 病 例', weight='normal', size=12, labelpad=6)

    ###中文字型###
    plt.rcParams['font.sans-serif'] = ['Taipei Sans TC Beta']  # 字型
    ##############

    plt.tight_layout()
    plt.savefig(PLOT7_path)
    
    #plt.show()
    
    plt.clf()
    time7_x.clear()  # 7天 => 在計算日期的時候直接存進去
    day7_new_number.clear()  # 存每日新增 # 7天 => 每天數值都存進去(會有7筆資料)

    ''' DAY 30 圖表製作 '''

    # 圖表變數設定
    time_x = []  # x軸計算用標#內有30天的所有日期
    time_real_x = []  # x軸實際的標#圖表中出現的五個日期

    day_new_number = []  # 存每日新增
    # 30天 => 每天數值都存進去

    # 30天，不包括今天(可以討論要不要改當天的，我看google的是昨天ㄉ)

    today = datetime.datetime.today()  # 直接取今天時間

    # 算日期
    # 用迴圈跑7次，倒算日期
    for i in range(30):

        time_x.insert(0, str(today.strftime("%m-%d")))  # 插入list的最前面，因為是倒算

        # 處理實際出現在x軸的日期
        if (i == 0) | (i == 6) | (i == 12) | (i == 18) | (i == 24):
            time_real_x.insert(0, str(today.strftime("%m-%d")))

        today -= datetime.timedelta(days=1)  # 減一天，最後一次會減到，但是不會被存進去
    #

    # 取數值
    for i in time_x:

        date = i

        # 確認是否有檔案
        try:  # 有，即進檔案抓資料

            file_position = COVID_data_path+'2021-'+date+'.xlsx'  # 轉檔案名字
            workbook = xlrd.open_workbook(file_position)

            # 選取sheet
            booksheet = workbook.sheet_by_name(country[j])  # 或用名稱取sheet

            #取daynew值 ###要換取其他值改這裡###
            sum_number = int(booksheet.cell_value(1, 2).replace(',', ''))  # 取day_new的值，去除分位號(replace)，轉數值

            # 加入y軸串列(直接加在最後)
            day_new_number.append(sum_number)
        except FileNotFoundError:
            sum_number = 0
            day_new_number.append(sum_number)

    # 30天製圖表

    plt.plot(time_x, day_new_number)
    plt.xticks(range(5, 31, 6), labels=time_real_x)

    plt.title(country_chinese[j], weight='bold', size=16)  # 直接取國家名
    plt.xlabel('日 期', weight='normal', size=12, labelpad=5)
    plt.ylabel('新 增 病 例', weight='normal', size=12, labelpad=7)

    ###中文字型###
    plt.rcParams['font.sans-serif'] = ['Taipei Sans TC Beta']  # 字型
    ##############
    plt.tight_layout()
    plt.savefig(PLOT30_path)  # 存圖檔一定要在show前面

    ##清理動態
    plt.clf()
    time_x.clear()  # x軸計算用標#內有30天的所有日期
    time_real_x.clear()  # x軸實際的標#圖表中出現的五個日期
    day_new_number.clear()  # 存每日新增
    
    '''上傳圖片'''
    # 7天
    img_url = upload_to_imgur(PLOT7_path)  # 圖片的url
    country_DAY7_url[j] = img_url
    # 30天
    img_url = upload_to_imgur(PLOT30_path)  # 圖片的url
    country_DAY30_url[j] = img_url

'''新建今日url excel檔案 '''
workbook = openpyxl.Workbook()
for i in country:

    # 建立一個新sheet
    sheet = workbook.create_sheet(country[i])  # 英文國家名
    # 國家名字
    sheet.cell(row=1, column=1, value="name")
    sheet.cell(row=2, column=1, value=country_chinese[i])
    # 7天連結
    sheet.cell(row=1, column=2, value="day_7_url")
    sheet.cell(row=2, column=2, value=country_DAY7_url[i])
    # 30天連結
    sheet.cell(row=1, column=3, value="day_30_url")
    sheet.cell(row=2, column=3, value=country_DAY30_url[i])

    # 儲存

workbook.save(savePLOTxlsx_name)
